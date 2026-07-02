"""
build_excel.py — Business Transformation Workbook
   Vegetarian Tiffin Home Kitchen → Deliveroo / Uber Eats
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path("outputs")

NAVY="FF1A3C5E"; CORAL="FFC0392B"; AMBER="FFE67E22"; GREEN="FF27AE60"
SKY="FF2980B9"; OFFWHITE="FFFAFAFA"; LTGREY="FFECF0F1"; MIDGREY="FF95A5A6"
DGREY="FF2C3E50"; WHITE="FFFFFFFF"; RED_BG="FFFDEDEC"; AMBER_BG="FFFEF9E7"
GREEN_BG="FFEAFAF1"; BLUE_INPUT="FF0000FF"

def fill(h):  return PatternFill("solid", fgColor=h)
def font(sz=11,color="FF000000",bold=False,italic=False):
    return Font(name="Arial",size=sz,color=color,bold=bold,italic=italic)
thin=Side(style="thin",color="FFBDC3C7")
def ba(c): c.border=Border(top=thin,bottom=thin,left=thin,right=thin)

def write(ws,r,c,v,sz=11,color="FF000000",bold=False,italic=False,
          bg=None,align="center",wrap=False,nf=None):
    cell=ws.cell(row=r,column=c,value=v)
    cell.font=font(sz,color,bold,italic)
    cell.alignment=Alignment(horizontal=align,vertical="center",wrap_text=wrap)
    if bg: cell.fill=fill(bg)
    ba(cell)
    if nf: cell.number_format=nf
    return cell

def banner(ws,title,sub,cols=8):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=cols)
    c=ws.cell(row=1,column=1,value=title)
    c.font=font(15,WHITE,bold=True); c.fill=fill(NAVY)
    c.alignment=Alignment(horizontal="left",vertical="center",indent=2)
    ws.row_dimensions[1].height=34
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=cols)
    s=ws.cell(row=2,column=1,value=sub)
    s.font=font(9,DGREY,italic=True); s.fill=fill(LTGREY)
    s.alignment=Alignment(horizontal="left",vertical="center",indent=2)
    ws.row_dimensions[2].height=16

def sec(ws,row,text,cols=8,bg=NAVY):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=cols)
    c=ws.cell(row=row,column=1,value=text)
    c.font=font(10,WHITE,bold=True); c.fill=fill(bg)
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[row].height=20

def hdr(ws,row,headers,widths,bg=NAVY):
    for i,(h,w) in enumerate(zip(headers,widths),1):
        c=ws.cell(row=row,column=i,value=h)
        c.font=font(10,WHITE,bold=True); c.fill=fill(bg)
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        ba(c)
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[row].height=28

def sp(ws,row,cols=8):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=cols)
    ws.row_dimensions[row].height=8

wb = Workbook()

# ════════════════════════════════════════════════════════════════════
# SHEET 1 — SITUATION & KEY NUMBERS
# ════════════════════════════════════════════════════════════════════
ws1=wb.active; ws1.title="Situation & Numbers"
ws1.sheet_view.showGridLines=False
banner(ws1,"Business Transformation: Vegetarian Tiffin → Deliveroo & Uber Eats",
       "Client: home kitchen, 30 orders/day, WhatsApp-only · Analyst: Vaishnavi Bhor")

# KPIs
kpis=[("Current Daily\nOrders","30","WhatsApp only",CORAL),
      ("Weighted AOV","£7.33","Current menu mix",AMBER),
      ("Annual Revenue\n(est.)","£57K","260 trading days",SKY),
      ("Gross Margin\n(direct)","60%","Pre-platform",GREEN),
      ("Platform\nCommission","30%","Deliveroo (typical)",CORAL),
      ("Setup Cost\n(one-off)","£1,215","To go live on platforms",AMBER),
      ("Revenue\nPotential","£130K+","At 100 orders/day",GREEN),
      ("Days to Register\nFood Business","28","Minimum by law — start today",CORAL)]
for i,(lbl,val,note,col) in enumerate(kpis,1):
    ws1.column_dimensions[get_column_letter(i)].width=16
    c1=ws1.cell(row=4,column=i,value=lbl); c1.font=font(9,WHITE,bold=True)
    c1.fill=fill(col); c1.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws1.row_dimensions[4].height=28
    c2=ws1.cell(row=5,column=i,value=val); c2.font=font(15,col,bold=True)
    c2.fill=fill(OFFWHITE); c2.alignment=Alignment(horizontal="center",vertical="center")
    ws1.row_dimensions[5].height=34
    c3=ws1.cell(row=6,column=i,value=note); c3.font=font(8,MIDGREY,italic=True)
    c3.fill=fill(OFFWHITE); c3.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws1.row_dimensions[6].height=22

sp(ws1,7)
sec(ws1,8,"SITUATION — WHAT EXISTS TODAY AND WHAT NEEDS TO CHANGE")
sit=[("The operation","30 vegetarian tiffin orders per day. Menu changes daily. Strong loyal customer base built on WhatsApp. Quality appears strong — the business has retained customers without any marketing."),
     ("The problem","Entirely dependent on WhatsApp for orders, discovery, and payment. Cannot scale beyond personal network. No digital presence, no food business registration, no platform listings."),
     ("The opportunity","Deliveroo and Uber Eats provide immediate access to thousands of local customers who have never heard of this kitchen. Vegetarian-only is a clear niche — easier to position and discover on platform search."),
     ("The constraint","Platform listing requires: (1) food business registration with local authority, (2) a Food Hygiene Rating, (3) standardised fixed menu with allergens, (4) product photos. None of these are currently in place."),
     ("What this project delivers","End-to-end transformation roadmap: current state economics → platform pricing model → menu standardisation → compliance checklist → 90-day phased implementation plan.")]
for i,(lbl,desc) in enumerate(sit,9):
    ws1.row_dimensions[i].height=32
    bg=LTGREY if i%2==0 else OFFWHITE
    c1=ws1.cell(row=i,column=1,value=lbl); c1.font=font(10,NAVY,bold=True)
    c1.fill=fill(bg); c1.alignment=Alignment(horizontal="left",vertical="center",indent=1); ba(c1)
    ws1.merge_cells(start_row=i,start_column=2,end_row=i,end_column=8)
    c2=ws1.cell(row=i,column=2,value=desc); c2.font=font(10)
    c2.fill=fill(bg); c2.alignment=Alignment(horizontal="left",vertical="center",indent=1,wrap_text=True); ba(c2)

sp(ws1,14)
sec(ws1,15,"ANALYTICAL FRAMEWORK — 5 WORKSTREAMS")
ws_items=[("WS1","Current State Economics","What does the business look like today? Revenue, margin, capacity.","Sheet: Platform Pricing"),
          ("WS2","Platform Economics & Pricing","Do current prices survive 30% commission? What must platform prices be?","Sheet: Platform Pricing"),
          ("WS3","Menu Standardisation","How to transform a daily WhatsApp menu into a permanent platform-ready listing.","Sheet: Standardised Menu"),
          ("WS4","Compliance & Legal Setup","What does UK law require before any food business lists on a platform?","Sheet: Compliance"),
          ("WS5","90-Day Implementation Roadmap","Phased plan: week-by-week, cost, owner, dependency.","Sheet: Roadmap")]
hdr(ws1,16,["Workstream","Name","Core Question","Deliverable"],[10,24,48,24],bg=DGREY)
for i,(ws_id,name,q,deliv) in enumerate(ws_items,17):
    ws1.row_dimensions[i].height=26
    bg=LTGREY if i%2==0 else OFFWHITE
    write(ws1,i,1,ws_id,10,CORAL,bold=True,bg=bg)
    write(ws1,i,2,name,10,DGREY,bold=True,bg=bg,align="left")
    write(ws1,i,3,q,10,bg=bg,align="left")
    write(ws1,i,4,deliv,9,SKY,bg=bg,align="left",italic=True)


# ════════════════════════════════════════════════════════════════════
# SHEET 2 — PLATFORM PRICING
# ════════════════════════════════════════════════════════════════════
ws2=wb.create_sheet("Platform Pricing")
ws2.sheet_view.showGridLines=False
banner(ws2,"WS2: Platform Economics & Pricing Strategy",
       "Current WhatsApp prices cannot go onto platforms unchanged · 30% commission requires ~30% price uplift")

sec(ws2,4,"THE CORE PROBLEM — WHY CURRENT PRICES DON'T WORK ON PLATFORM")
ws2.row_dimensions[5].height=28
prob_text=("A £7 WhatsApp order on Deliveroo at 30% commission: Deliveroo takes £2.10. "
           "Restaurant receives £4.90. After food + packaging cost (40% of AOV = £2.80), gross profit is £2.10 — a 30% margin. "
           "That is marginally viable, BUT: (1) the customer also pays a £2.99 Deliveroo service fee on top, "
           "making the real customer cost £9.99. At that customer price point, a £7 restaurant price looks cheap and attracts "
           "value expectations that are hard to manage. (2) The right approach is to price the platform menu at £9-10 and "
           "maintain a comparable gross margin to the direct channel.")
ws2.merge_cells(start_row=5,start_column=1,end_row=5,end_column=8)
c=ws2.cell(row=5,column=1,value=prob_text); c.font=font(10)
c.fill=fill(AMBER_BG); c.alignment=Alignment(horizontal="left",vertical="center",indent=1,wrap_text=True); ba(c)

sp(ws2,6)
sec(ws2,7,"UNIT ECONOMICS: EACH MENU ITEM × EACH CHANNEL")
hdr(ws2,8,["Menu Item","WhatsApp\nPrice","Platform\nPrice (+30%)","Deliveroo\nCommission (30%)","Restaurant\nReceives","Ingredient\nCOGS (40%)","Gross\nProfit","Gross\nMargin"],
    [28,14,16,20,18,18,14,14])

menu_pricing=[
    ("Tiffin Combo Regular",         7.00,  9.49, 0.30),
    ("Tiffin Combo Large",           7.00, 10.49, 0.30),
    ("Paneer Bhurji Solo (Large)",   8.00, 10.99, 0.30),
    ("Masoor Dal Solo (Large)",      7.00,  9.99, 0.30),
    ("Family Sharing Box",          14.00, 17.99, 0.30),
    ("Today's Special",              6.00,  8.49, 0.30),
]
for i,(name,wp,pp,cr) in enumerate(menu_pricing,9):
    ws2.row_dimensions[i].height=26
    bg=LTGREY if i%2==0 else OFFWHITE
    commission=pp*cr; receives=pp-commission; cogs=wp*0.40; gp=receives-cogs; gm=gp/pp
    gp_col=GREEN if gm>0.30 else AMBER if gm>0.15 else CORAL
    write(ws2,i,1,name,10,DGREY,bold=True,bg=bg,align="left")
    write(ws2,i,2,wp,10,BLUE_INPUT,bg=bg,nf="£#,##0.00")
    write(ws2,i,3,pp,10,GREEN,bold=True,bg=bg,nf="£#,##0.00")
    write(ws2,i,4,commission,10,CORAL,bg=bg,nf="£#,##0.00")
    write(ws2,i,5,receives,10,bg=bg,nf="£#,##0.00")
    write(ws2,i,6,cogs,10,CORAL,bg=bg,nf="£#,##0.00")
    write(ws2,i,7,gp,11,gp_col,bold=True,bg=bg,nf="£#,##0.00")
    write(ws2,i,8,gm,11,gp_col,bold=True,bg=bg,nf="0.0%")

sp(ws2,15)
sec(ws2,16,"CHANNEL COMPARISON — WHAT THE RESTAURANT KEEPS ON A £7 EQUIVALENT SALE")
hdr(ws2,17,["Channel","Customer Pays","Restaurant Receives","COGS","Gross Profit","Gross Margin","Notes"],
    [22,16,20,12,14,14,30])
channel_data=[
    ("WhatsApp (direct)",       "£7.00 + self-delivery",    7.00,  2.80, 4.20, 0.60, "No commission. Best margin. Limited reach."),
    ("Deliveroo (30%)",         "£9.00 + £2.99 svc fee",    6.30,  2.80, 3.50, 0.39, "Use for discovery. Price 30% above WhatsApp."),
    ("Uber Eats Standard (25%)","£9.00 + svc fee",           6.75,  2.80, 3.95, 0.44, "Better margin than Deliveroo. Good starting point."),
    ("Own Website (Stripe 3%)", "£7.70 (10% above direct)", 7.47,  2.80, 4.67, 0.61, "Best long-term channel. Lower commission. Own data."),
]
for i,(ch,cust,rec,cogs,gp,gm,note) in enumerate(channel_data,18):
    ws2.row_dimensions[i].height=24
    bg=LTGREY if i%2==0 else OFFWHITE
    gp_col=GREEN if gm>0.50 else AMBER if gm>0.35 else CORAL
    write(ws2,i,1,ch,10,DGREY,bold=True,bg=bg,align="left")
    write(ws2,i,2,cust,10,bg=bg)
    write(ws2,i,3,rec,10,bg=bg,nf="£#,##0.00")
    write(ws2,i,4,cogs,10,CORAL,bg=bg,nf="£#,##0.00")
    write(ws2,i,5,gp,11,gp_col,bold=True,bg=bg,nf="£#,##0.00")
    write(ws2,i,6,gm,11,gp_col,bold=True,bg=bg,nf="0.0%")
    write(ws2,i,7,note,9,DGREY,bg=bg,align="left",italic=True)


# ════════════════════════════════════════════════════════════════════
# SHEET 3 — STANDARDISED MENU
# ════════════════════════════════════════════════════════════════════
ws3=wb.create_sheet("Standardised Menu")
ws3.sheet_view.showGridLines=False
banner(ws3,"WS3: Standardised Platform Menu — WhatsApp Menu → Platform-Ready Listing",
       "Platforms require: fixed permanent items · gram weights · allergen info (LEGAL) · product photos")

sec(ws3,4,"WHY THE CURRENT MENU CANNOT GO ON DELIVEROO/UBER EATS AS-IS")
problems=[("Daily changing menu","Platforms need a permanent listing. You cannot update your full menu every day. A rotating 'Today's Special' slot is the right solution for daily variety.","BLOCKING"),
          ("No allergen information","UK law requires allergen info for all 14 allergens on every menu item sold for delivery. Deliveroo will not list you without this. Failing to provide it is a criminal offence.","LEGAL — BLOCKING"),
          ("No gram weights or portion clarity","'Medium' and 'Large' mean nothing to a new customer who has never ordered before. Platform listings need: '250g basmati rice', '2 chapatis'.","CONVERSION — BLOCKING"),
          ("No product photos","Restaurants with photos get 40%+ more clicks on Deliveroo and Uber Eats (both platforms publish this data). A listing without photos will not convert.","CONVERSION — CRITICAL"),
          ("8 options per day","Too many choices cause decision paralysis. Best practice for a small kitchen: 5-6 permanent items + 1 daily special. Reduces kitchen complexity too.","OPERATIONAL"),]
hdr(ws3,5,["Problem","Detail","Severity"],[28,52,20],bg=CORAL)
for i,(p,d,sev) in enumerate(problems,6):
    ws3.row_dimensions[i].height=32
    bg=RED_BG if "BLOCKING" in sev or "LEGAL" in sev else AMBER_BG
    sev_col=CORAL if "BLOCKING" in sev or "LEGAL" in sev else AMBER
    write(ws3,i,1,p,10,CORAL,bold=True,bg=bg,align="left")
    write(ws3,i,2,d,10,bg=bg,align="left",wrap=True)
    write(ws3,i,3,sev,9,sev_col,bold=True,bg=bg)

sp(ws3,11)
sec(ws3,12,"RECOMMENDED PLATFORM MENU — 6 PERMANENT ITEMS")
hdr(ws3,13,["Item","Platform Description","WhatsApp\nPrice","Platform\nPrice","Portion\nSize","Allergens\n(14-allergen check)","Photo\nPriority","Est.\nOrders/Day"],
    [22,40,12,14,14,34,14,14])

std_menu=[
    ("Tiffin Combo\n(Regular)","Paneer Bhurji or Masoor Dal + Basmati Rice (250g) + 3 Chapatis. Home-style North Indian vegetarian, freshly made daily.",7.00,9.49,"~500g","Gluten (chapati), Dairy (paneer). May contain nuts.","HIGH",12),
    ("Tiffin Combo\n(Large)","Paneer Bhurji or Masoor Dal + Basmati Rice (350g) + 4 Chapatis. Bigger appetite? This is the one.",7.00,10.49,"~700g","Gluten (chapati), Dairy (paneer). May contain nuts.","HIGH",8),
    ("Paneer Bhurji\n(Solo, Large)","Rich Paneer Peas Bhurji in tomato-onion masala. With 4 chapatis OR large basmati rice — choose at checkout.",8.00,10.99,"~400g","Dairy (paneer), Gluten if chapati. May contain nuts.","HIGH",4),
    ("Masoor Dal\n(Solo, Large)","Red lentil dal with whole spices, cooked slow. With 4 chapatis OR large basmati rice — choose at checkout. Fully vegan without chapati.",7.00,9.99,"~400g","Gluten if chapati selected. Vegan option available. May contain nuts.","HIGH",3),
    ("Family Sharing\nBox","Large Bhurji + Large Masoor + 6 chapatis + large rice. Feeds 2-3 people. The full home-cooked meal.",14.00,17.99,"~1.2kg","Gluten (chapati), Dairy (paneer). May contain nuts.","HIGH",2),
    ("Today's\nSpecial","Ask us what's cooking! Daily rotating dish — seasonal vegetables, extra curries, or festive specials. Allergens listed daily in description.",6.00,8.49,"Varies","Listed daily — check item description.","LOW",1),
]
for i,(name,desc,wp,pp,portion,allergen,photo,orders) in enumerate(std_menu,14):
    ws3.row_dimensions[i].height=38
    bg=LTGREY if i%2==0 else OFFWHITE
    write(ws3,i,1,name,10,NAVY,bold=True,bg=bg,align="left",wrap=True)
    write(ws3,i,2,desc,9,bg=bg,align="left",wrap=True)
    write(ws3,i,3,wp,10,MIDGREY,bg=bg,nf="£#,##0.00")
    write(ws3,i,4,pp,10,GREEN,bold=True,bg=bg,nf="£#,##0.00")
    write(ws3,i,5,portion,9,bg=bg)
    write(ws3,i,6,allergen,9,CORAL,bg=bg,align="left",wrap=True,italic=True)
    write(ws3,i,7,photo,9,GREEN if photo=="HIGH" else AMBER,bold=True,bg=bg)
    write(ws3,i,8,orders,10,bg=bg)


# ════════════════════════════════════════════════════════════════════
# SHEET 4 — COMPLIANCE CHECKLIST
# ════════════════════════════════════════════════════════════════════
ws4=wb.create_sheet("Compliance Checklist")
ws4.sheet_view.showGridLines=False
banner(ws4,"WS4: UK Compliance & Legal Setup — Complete Before Going Live",
       "Sources: FSA (food.gov.uk) · GOV.UK · Deliveroo (Delta Digital 2025) · Uber Eats (Deliverect 2025)")

sec(ws4,4,"THE LEGAL PICTURE — WHAT UK LAW REQUIRES OF A HOME FOOD BUSINESS")
legal_context=[("Who this applies to","Any UK business that regularly supplies food to the public — even from a home kitchen. 'Regularly and organised basis' = you are a food business. WhatsApp tiffin orders = food business."),
               ("Registration","FREE and mandatory. Must be done at least 28 days before trading (or immediately if already trading). Cannot be refused. Apply via your local council website."),
               ("Food hygiene certificate","Not compulsory but FSA strongly recommends it. Level 2 Food Safety for Catering. Online course, ~£25, one day. Deliveroo/Uber Eats may require it."),
               ("HMRC self-employment","Required — even if part-time. Register at gov.uk/register-for-self-assessment. Tax on profits above personal allowance."),
               ("Allergen law","UK Food Information for Consumers Regulation. 14 allergens must be declared on every dish sold for delivery. Failure = criminal offence. Dairy and gluten both present in this menu."),
               ("Food Hygiene Rating Scheme","Inspected by Environmental Health after registration. Ratings 0-5. Deliveroo requires minimum rating to list. Aim for 5 — all businesses should be able to achieve this."),]
hdr(ws4,5,["Legal Item","What It Means in Practice"],[24,66],bg=CORAL)
for i,(item,detail) in enumerate(legal_context,6):
    ws4.row_dimensions[i].height=30
    bg=RED_BG if i%2==0 else AMBER_BG
    write(ws4,i,1,item,10,CORAL,bold=True,bg=bg,align="left")
    ws4.merge_cells(start_row=i,start_column=2,end_row=i,end_column=8)
    c=ws4.cell(row=i,column=2,value=detail); c.font=font(10)
    c.fill=fill(bg); c.alignment=Alignment(horizontal="left",vertical="center",indent=1,wrap_text=True); ba(c)

sp(ws4,12)
sec(ws4,13,"FULL COMPLIANCE & SETUP CHECKLIST — TICK BEFORE GOING LIVE")
hdr(ws4,14,["Item","Category","Cost (£)","Timeline","Status","Action Required","Source"],
    [38,14,12,18,14,28,26])

checklist=[
    ("Register food business with local authority","LEGAL",0,"28 days before trading","⬜ TODO","Go to local council website now — search '[your council] food business registration'","FSA / GOV.UK"),
    ("Register as self-employed with HMRC","LEGAL",0,"Before trading","⬜ TODO","gov.uk/register-for-self-assessment","GOV.UK / HMRC"),
    ("Level 2 Food Hygiene Certificate (online)","LEGAL",25,"1 day (online)","⬜ TODO","highspeedtraining.co.uk or similar — ~£25","FSA guidance"),
    ("Document allergens for all 6 menu items","LEGAL",0,"Immediate","⬜ TODO","Use FSA allergen checker tool. List all 14 allergens per item.","UK FIC Regulation"),
    ("HACCP-lite food safety system","LEGAL",0,"1-2 days","⬜ TODO","Download FSA 'Safer Food Better Business' pack — free","FSA"),
    ("Public liability insurance","RECOMMENDED",150,"1-2 days","⬜ TODO","Compare on Simply Business or PolicyBee — ~£150/yr","Industry standard"),
    ("Food-grade packaging with allergen labelling","LEGAL",100,"Before go-live","⬜ TODO","Labels must include: business name, allergens, storage instructions","UK food labelling law"),
    ("Environmental Health kitchen inspection","LEGAL",0,"Council arranges post-registration","⬜ AUTO","Happens automatically after registration — make kitchen inspection-ready","FSA"),
    ("5-star Food Hygiene Rating","PLATFORM",0,"At inspection","⬜ TODO","Clean and document everything. Use FSA checklist to self-audit first.","FHRS"),
    ("Deliveroo listing application","PLATFORM",510,"After 5-star rating","⬜ TODO","Apply at restaurants.deliveroo.com — £510 onboarding fee (8 instalments)","Delta Digital 2025"),
    ("Uber Eats listing application","PLATFORM",0,"After food hygiene cert","⬜ TODO","Apply at merchants.ubereats.com — faster onboarding than Deliveroo","Deliverect 2025"),
    ("Product photography (6 menu items)","PLATFORM",200,"1 day shoot","⬜ TODO","Hire local food photographer OR use Deliveroo's in-house photo service","Platform conversion data"),
    ("Google Business Profile setup","RECOMMENDED",0,"2 hours","⬜ TODO","business.google.com — free, drives local organic discovery","Industry standard"),
    ("Simple website with menu + ordering link","RECOMMENDED",200,"1-2 weeks","⬜ TODO","Squarespace or Wix — £15/month. Embed Uber Eats/Deliveroo order link.","Recommended"),
    ("Instagram page with food photos","RECOMMENDED",0,"1-2 hours","⬜ TODO","Username = business name. Link in bio to delivery platform.","Industry standard"),
]
urgency_bg={"LEGAL":RED_BG,"PLATFORM":AMBER_BG,"RECOMMENDED":OFFWHITE}
for i,(item,cat,cost,timeline,status,action,source) in enumerate(checklist,15):
    ws4.row_dimensions[i].height=28
    bg=urgency_bg.get(cat,OFFWHITE)
    cat_col=CORAL if cat=="LEGAL" else AMBER if cat=="PLATFORM" else SKY
    write(ws4,i,1,item,10,DGREY,bg=bg,align="left",wrap=True)
    write(ws4,i,2,cat,9,cat_col,bold=True,bg=bg)
    write(ws4,i,3,cost,10,BLUE_INPUT if cost>0 else DGREY,bg=bg,nf="£#,##0" if cost>0 else "@")
    write(ws4,i,4,timeline,9,bg=bg)
    write(ws4,i,5,status,10,bg=bg)
    write(ws4,i,6,action,9,DGREY,bg=bg,align="left",wrap=True)
    write(ws4,i,7,source,8,MIDGREY,bg=bg,align="left",italic=True)

# Total row
ws4.row_dimensions[30].height=24
total_cost=sum(c for _,_,c,*_ in checklist)
for col,(v,nf) in enumerate([("TOTAL SETUP COST",None),(None,None),(total_cost,"£#,##0"),(None,None),(None,None),(None,None),(None,None)],1):
    cell=ws4.cell(row=30,column=col,value=v)
    cell.font=font(11,WHITE,bold=True); cell.fill=fill(NAVY)
    ba(cell); cell.alignment=Alignment(horizontal="center",vertical="center")
    if nf and v: cell.number_format=nf


# ════════════════════════════════════════════════════════════════════
# SHEET 5 — 90-DAY ROADMAP
# ════════════════════════════════════════════════════════════════════
ws5=wb.create_sheet("90-Day Roadmap")
ws5.sheet_view.showGridLines=False
banner(ws5,"WS5: 90-Day Implementation Roadmap — Phase by Phase",
       "Phase 1: Legal foundation (Weeks 1-4) · Phase 2: Platform launch (Weeks 5-8) · Phase 3: Scale (Weeks 8-12)")

sec(ws5,4,"PHASE 1 — LEGAL FOUNDATION (WEEKS 1–4): DO NOT SKIP THIS")
ws5.row_dimensions[5].height=24
ws5.merge_cells(start_row=5,start_column=1,end_row=5,end_column=8)
c=ws5.cell(row=5,column=1,value="⚠  CRITICAL: The food business registration takes 28 days minimum before you can legally trade. Start this on Day 1, regardless of anything else. Everything else in this roadmap depends on it.")
c.font=font(10,CORAL,bold=True); c.fill=fill(RED_BG)
c.alignment=Alignment(horizontal="left",vertical="center",indent=1); ba(c)

hdr(ws5,6,["Week","Task","Owner","Cost","Dependency","Why This Week"],
    [10,44,12,10,24,26])

phase1=[("Week 1","Register food business with local council (online, 10 minutes)","Founder","£0","None — do today","28-day clock starts now. Everything is blocked until this is done."),
        ("Week 1","Register as self-employed with HMRC (online, 20 minutes)","Founder","£0","None","Tax compliance — HMRC tracks Deliveroo/Uber Eats income from Jan 2024."),
        ("Week 1","Complete Level 2 Food Hygiene Certificate online (1 day)","Founder","£25","None","Required for inspection, platforms prefer it. Cheap and fast."),
        ("Week 1-2","Document allergens for all 6 standardised menu items","Founder","£0","Menu finalised","LEGAL REQUIREMENT. Use FSA allergen checker. Keep written records."),
        ("Week 2","Download and complete FSA 'Safer Food Better Business' pack","Founder","£0","Food hygiene cert","Written food safety system = council will want to see this at inspection."),
        ("Week 2-3","Finalise standardised 6-item menu with descriptions and gram weights","Founder","£0","None","Platform needs fixed menu to submit listing application."),
        ("Week 3","Purchase public liability insurance","Founder","£150","Business registered","Protects against delivery-related claims. Required before operating at scale."),
        ("Week 3-4","Source food-grade packaging with allergen/business name labels","Founder","£100","Menu finalised","Legal requirement for delivery orders. Order in bulk — cost per unit drops."),]

for i,(wk,task,owner,cost,dep,why) in enumerate(phase1,7):
    ws5.row_dimensions[i].height=30
    bg=RED_BG if i%2==0 else OFFWHITE
    write(ws5,i,1,wk,10,CORAL,bold=True,bg=bg)
    write(ws5,i,2,task,10,bg=bg,align="left",wrap=True)
    write(ws5,i,3,owner,9,DGREY,bg=bg)
    write(ws5,i,4,cost,10,BLUE_INPUT if "£" in cost and cost!="£0" else MIDGREY,bg=bg)
    write(ws5,i,5,dep,9,DGREY,bg=bg,align="left",italic=True)
    write(ws5,i,6,why,9,CORAL,bg=bg,align="left",wrap=True)

sp(ws5,15)
sec(ws5,16,"PHASE 2 — PLATFORM LAUNCH (WEEKS 5–8)",bg=AMBER)

phase2=[("Week 4-5","Environmental Health kitchen inspection (council arranges)","Council","£0","Registration complete","If kitchen is clean and documented — aim for 5 stars immediately."),
        ("Week 5","Professional food photography (6 items + setup shot)","Photographer","£200","Menu finalised","40%+ conversion uplift. Non-negotiable for platform listings. Book in advance."),
        ("Week 5-6","Apply to Uber Eats — upload menu, photos, pricing","Founder","£0","5-star rating + photos","Faster onboarding than Deliveroo. Start here. Go live quickly."),
        ("Week 5-6","Apply to Deliveroo — upload menu, photos, pricing","Founder","£510","5-star rating + photos","£510 spread over 8 payment deductions from earnings. Larger audience."),
        ("Week 6","Set pricing 30% above WhatsApp prices on both platforms","Founder","£0","Platform accounts live","Standard industry practice. Absorbs commission. Keeps margin comparable."),
        ("Week 6-7","Build simple website: Squarespace or Wix with order links","Founder","£200","None — run in parallel","Own digital presence = not fully dependent on platform algorithms."),
        ("Week 7","Set up Google Business profile with photos and menu","Founder","£0","Website live","Free local search visibility. Customers search 'vegetarian food delivery [area]'."),
        ("Week 7-8","Soft launch: tell WhatsApp customers about Deliveroo listing, ask for reviews","Founder","£0","Platform listing live","First 10 reviews are critical for platform algorithm placement."),]

hdr(ws5,17,["Week","Task","Owner","Cost","Dependency","Why This Week"],
    [10,44,12,10,24,26],bg=AMBER)
for i,(wk,task,owner,cost,dep,why) in enumerate(phase2,18):
    ws5.row_dimensions[i].height=30
    bg=AMBER_BG if i%2==0 else OFFWHITE
    write(ws5,i,1,wk,10,AMBER,bold=True,bg=bg)
    write(ws5,i,2,task,10,bg=bg,align="left",wrap=True)
    write(ws5,i,3,owner,9,DGREY,bg=bg)
    write(ws5,i,4,cost,10,BLUE_INPUT if "£" in cost and cost not in ["£0","£510"] else MIDGREY,bg=bg)
    write(ws5,i,5,dep,9,DGREY,bg=bg,align="left",italic=True)
    write(ws5,i,6,why,9,AMBER,bg=bg,align="left",wrap=True)

sp(ws5,26)
sec(ws5,27,"PHASE 3 — SCALE (WEEKS 9–12)",bg=GREEN)

phase3=[("Week 8-10","Monitor platform: acceptance rate, avg rating, prep time compliance","Founder","£0","Both platforms live","Low acceptance rate or slow prep = platform penalises your visibility."),
        ("Week 9-10","Add 'Today's Special' as a rotating daily item on platform","Founder","£0","Base menu stable","Brings back repeat customers. Easy to manage once workflow is set."),
        ("Week 10-11","Instagram: post 3x/week — food photos, daily special, behind-scenes","Founder","£0","Photos available","Vegetarian food photographs well. Build following before paid ads."),
        ("Week 11-12","Review P&L: direct vs platform channel split. Adjust if needed","Founder","£0","Full month of platform data","Data from first full month: which platform performs, which items sell."),
        ("Week 12","Assess kitchen capacity for 75-100 order target. Plan next hire if needed","Founder","£0","Volume trending up","At 75+ orders/day, solo operation becomes difficult. Plan ahead."),]

hdr(ws5,28,["Week","Task","Owner","Cost","Dependency","Focus"],
    [10,44,12,10,24,26],bg=GREEN)
for i,(wk,task,owner,cost,dep,focus) in enumerate(phase3,29):
    ws5.row_dimensions[i].height=30
    bg=GREEN_BG if i%2==0 else OFFWHITE
    write(ws5,i,1,wk,10,GREEN,bold=True,bg=bg)
    write(ws5,i,2,task,10,bg=bg,align="left",wrap=True)
    write(ws5,i,3,owner,9,DGREY,bg=bg)
    write(ws5,i,4,cost,10,MIDGREY,bg=bg)
    write(ws5,i,5,dep,9,DGREY,bg=bg,align="left",italic=True)
    write(ws5,i,6,focus,9,GREEN,bg=bg,align="left",wrap=True)


# ════════════════════════════════════════════════════════════════════
# SHEET 6 — RECOMMENDATIONS
# ════════════════════════════════════════════════════════════════════
ws6=wb.create_sheet("Recommendations")
ws6.sheet_view.showGridLines=False
banner(ws6,"Key Recommendations & Honest Assessment",
       "What to do, in what order, and what to expect — honestly")

sec(ws6,4,"THE FIVE THINGS THAT MATTER MOST")
rec_rows=[
    ("1 — LEGAL FIRST","Register as a food business TODAY. The 28-day waiting period before trading is non-negotiable. If you are already trading without registration, this is urgent. Everything else in this plan cannot happen until registration is complete and you have a Food Hygiene Rating.","URGENT — do today"),
    ("2 — REPRICE FOR PLATFORMS","Your current £6-8 menu cannot go on Deliveroo/Uber Eats unchanged. At 30% commission, a £7 item leaves only 30% gross margin — workable, but the customer also pays £2.99 service fee, making their real cost £9.99. Price transparently: set platform prices at £9-10. This is standard industry practice and no customer will complain if the quality is there.","Price before listing"),
    ("3 — FIX THE MENU BEFORE APPLYING","Deliveroo and Uber Eats applications require allergen information for every item. You need a standardised, permanent menu with descriptions, gram weights, and allergen declarations. Consolidate from 8 daily options to 6 permanent items plus one daily special. This simplifies your kitchen workflow too.","Before platform apply"),
    ("4 — PHOTOGRAPHS ARE NON-NEGOTIABLE","Both platforms publish data showing that listings with professional food photos receive 40%+ more clicks. A £200 food photography session (6 items) is the highest-ROI spend in this entire plan. Without photos, your listing will not convert — it will just sit there.","Budget £200 for this"),
    ("5 — THINK HYBRID FROM DAY ONE","Do not become 100% dependent on Deliveroo/Uber Eats. They take 30% of every order. Your loyal WhatsApp customers should eventually order via your own website (3% Stripe fee). Use the platforms for discovery — they are your marketing channel, not your business model. Build your own customer list from day one.","Long-term strategy"),
]
hdr(ws6,5,["Priority","Recommendation","Timing"],[14,62,20])
for i,(pri,rec,timing) in enumerate(rec_rows,6):
    ws6.row_dimensions[i].height=42
    bg=RED_BG if i==6 else AMBER_BG if i<8 else OFFWHITE
    pri_col=CORAL if i<=7 else AMBER
    write(ws6,i,1,pri,10,pri_col,bold=True,bg=bg)
    write(ws6,i,2,rec,10,bg=bg,align="left",wrap=True)
    write(ws6,i,3,timing,9,GREEN,bold=True,bg=bg,wrap=True)

sp(ws6,11)
sec(ws6,12,"HONEST ASSESSMENT — WHAT IS REALISTIC")
honest=[("Revenue projection","At 100 orders/day (50 WhatsApp + 50 platform), after commission: ~£130K annual net revenue vs ~£57K today. This assumes the platforms perform — it is not guaranteed. In the first 3 months, expect lower volumes as the listing builds reviews and algorithm placement."),
        ("The platform algorithm","Both Deliveroo and Uber Eats prioritise restaurants with: high acceptance rate (>95%), fast prep time, high rating (>4.5 stars), and good review volume. In the first 30 days, your placement will be low. This improves with volume and ratings. Mobilising WhatsApp customers to leave Deliveroo reviews immediately is critical."),
        ("Kitchen capacity","At 30 orders/day, one person can manage. At 75+ orders/day with platform volume, the timing pressure increases significantly — platform orders arrive in clusters and have defined prep time expectations. Plan your kitchen workflow before hitting this volume, not after."),
        ("Deliveroo vs Uber Eats","Start with Uber Eats first — no onboarding fee, faster approval, good volume. Add Deliveroo after 4-6 weeks — larger UK audience, but the £510 onboarding cost and slower process make it a second step."),
        ("What could go wrong","Poor Food Hygiene Rating (below 3) = cannot list. Low review score = platform suppresses your listing. Slow prep times = orders cancelled and rating drops. HMRC reporting = your Deliveroo/Uber Eats income is automatically reported to HMRC from 2024. Declare it."),]
hdr(ws6,13,["Area","Honest Assessment"],[20,70],bg=DGREY)
for i,(area,detail) in enumerate(honest,14):
    ws6.row_dimensions[i].height=36
    bg=LTGREY if i%2==0 else OFFWHITE
    write(ws6,i,1,area,10,NAVY,bold=True,bg=bg,align="left")
    ws6.merge_cells(start_row=i,start_column=2,end_row=i,end_column=8)
    c=ws6.cell(row=i,column=2,value=detail); c.font=font(10)
    c.fill=fill(bg); c.alignment=Alignment(horizontal="left",vertical="center",indent=1,wrap_text=True); ba(c)

sp(ws6,19)
ws6.merge_cells("A20:H20")
cf=ws6.cell(row=20,column=1,value="Vaishnavi Bhor · MSc Business Analytics, University of Manchester · vbhor207@gmail.com · linkedin.com/in/vaishnavi-bhor-business-analyst · vbho.github.io/portfolio")
cf.font=font(10,SKY); cf.fill=fill(OFFWHITE); cf.alignment=Alignment(horizontal="left",vertical="center",indent=1); ba(cf)
ws6.row_dimensions[20].height=20

outpath=OUT/"Tiffin_Business_Transformation_VaishnaviBhor.xlsx"
wb.save(outpath)
print(f"\n✓  Workbook saved → {outpath}")
print(f"   Sheets: {len(wb.sheetnames)}")
